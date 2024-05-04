from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup
from glob import glob
from nltk.tokenize import WordPunctTokenizer, PunktSentenceTokenizer
from nltk.corpus import stopwords
from string import punctuation
from tqdm import tqdm
import warnings
warnings.filterwarnings('ignore')

# taking input from Input.xlsx
wb_input = load_workbook('Input.xlsx')
ws_input = wb_input['Sheet1']

# reading rows
for row in tqdm(range(2, 102), colour='green'):
    urlId = ws_input.cell(row, 1).value
    filepath = f'extracted_text/{urlId}.txt'

    # storing text in a text file
    with open(filepath,'w', encoding='utf-8') as file:
        link = ws_input.cell(row, 2).value
        result = requests.get(link)
        if result.status_code==200:
            doc = BeautifulSoup(result.text, 'html.parser')
            if doc.find('h1', {'class':'entry-title'}):
                title = doc.find('h1', {'class':'entry-title'})
                file.write(title.get_text() + '\n')

            # for some webpages have the class tdb-title-text
            elif doc.find('h1', {'class':'tdb-title-text'}):
                title = doc.find('h1', {'class':'tdb-title-text'})
                file.write(title.get_text() + '\n')
            text = doc.find('div', {'class':'td-post-content'}).find_all(['p','li'])
            for element in text:
                file.write(element.get_text() + '\n')
        else:
            print(f'Page Not Found with error code {result.status_code}')
            file.write('page not found')

wb_input.close()

# # text analysis - extracting provided folders
stopwords_files = sorted(glob('StopWords/*.txt'))
masterdictionary_files = sorted(glob('MasterDictionary/*.txt'))

stopwords_list = []
for file in stopwords_files:
    with open(file, 'r', encoding='latin1') as file:
        stopwords_list.append(file.read().splitlines())

masterdictionary_list = []
for file in masterdictionary_files:
    with open(file, 'r', encoding='latin1') as file:
        masterdictionary_list.append(file.read().splitlines())

stopword_auditor, stopword_currencies, stopword_dateandnumbers, stopword_generic, stopword_genericlong, stopword_geographic, stopword_names = stopwords_list

masterdictionary_negative, masterdictionary_positive = masterdictionary_list
masterdictionary_positive, masterdictionary_negative = set(masterdictionary_positive), set(masterdictionary_negative)

# cleaning the words in the files with pipe '|'
pipe_contained = [stopword_currencies, stopword_dateandnumbers, stopword_geographic, stopword_names]

n = len(pipe_contained)
for i in range(n):
    pipe_contained[i] = [c.split('|')[0].strip() for c in pipe_contained[i]]

stopword_currencies, stopword_dateandnumbers, stopword_geographic, stopword_names = pipe_contained



# reading extracted file
extracted_files = sorted(glob('extracted_text/*.txt'))
custom_stopword_list = [stopword_auditor, stopword_currencies, stopword_dateandnumbers, stopword_generic, stopword_genericlong, stopword_geographic, stopword_names]

wb_output = load_workbook('Input.xlsx')
ws_output = wb_output['Sheet1']

# get cleaned words - getting stopwords from folders is due
def getWordDetails(file):
    file.seek(0)
    tk = WordPunctTokenizer()
    data = file.read()
    punctuations = set(punctuation)
   
    tokens = tk.tokenize(data)
    punctuation_free = []
    for token in tokens:
        punctuation_free.append(''.join(c for c in token if c not in punctuations))
    punctuation_free = [words for words in punctuation_free if words != '']

    stopwords_free = punctuation_free
    for custom_stopwords in custom_stopword_list:
        stopwords_free = [words for words in stopwords_free if words not in custom_stopwords]

    total_word_count = len(punctuation_free)
    customsw_word_count = len(stopwords_free)

    return punctuation_free, stopwords_free, customsw_word_count, total_word_count

# get sentence count
def getSentenceCount(file):
    file.seek(0)
    tk = PunktSentenceTokenizer()
    return len(tk.tokenize(file.read()))
    
# get Positive and Negative Score
def getPositiveNegativeScore(words_list):
    positive_count, negative_count = 0, 0
    for words in words_list:
        if words in masterdictionary_positive:
            positive_count += 1
        
        elif words in masterdictionary_negative:
            negative_count -= 1
    return positive_count, negative_count*-1

# get Polar Score
def getPolarScore(positive_count, negative_count):
    return (positive_count - negative_count)/ ((positive_count + negative_count) + 0.000001)

# get Subjectivity Score
def getSubjectivityScore(positive_count, negative_count, word_count):
    return (positive_count + negative_count)/ ((word_count) + 0.000001)

# get syllable count to check for complex words
def getSyllableCountWord(word):
    vowels = 'aeiou'
    c = sum(1 for ch in word if ch in vowels)
    if word.endswith(('es', 'ed')):
        return c-1
    return c

# count Complex words
def getComplexWordCount(words_list):
    c = 0
    for word in words_list:
        if getSyllableCountWord(word) > 2:
            c += 1
    return c

# count personal pronouns
def getPersonalPronouns(words_list):
    pronouns = ['I', 'we', 'my', 'ours', 'us', 'We', 'My', 'Ours', 'Us']
    c = 0
    for word in words_list:
        if word in pronouns:
            c += 1
    return c

# get average word length
def getAvgWordLength(words_list, total_word_count):
    c = 0
    for word in words_list:
        c += sum(1 for ch in word)
    return c/total_word_count

# As the "Syllable Count Per Word" is not mentioned whether it needed to sum all counts of syllables of the words in the text or needed to return as a dict or a sequence of list,
# or any particular formula should be implemented
# Here, I am returning the total amount collectively.
def countSyllables(words_list):
    return sum(getSyllableCountWord(word) for word in words_list)





print('Starting Text Analysis')
for row, file in tqdm(enumerate(extracted_files), colour='green'):
    with open(file, 'r', encoding='utf-8') as file:
        punt_free_words, cleaned_words, cleaned_word_count, total_word_count = getWordDetails(file)
        sentence_count = getSentenceCount(file)
        nltk_stopwords = set(stopwords.words('english'))
        nltk_stopwords_free = [words for words in punt_free_words if words not in nltk_stopwords]
        nltk_free_wc = len(nltk_stopwords_free)

        # 1. sentiment analysis
        positive_score, negative_score = getPositiveNegativeScore(cleaned_words)
        polar_score = getPolarScore(positive_score, negative_score)
        subjective_score = getSubjectivityScore(positive_score, negative_score, cleaned_word_count)

        # 2. Analysis of readbility
        # 4. Complex word count
        avg_sentence_length = total_word_count/sentence_count
        count_complex_words = getComplexWordCount(punt_free_words)
        perc_complex_words = count_complex_words/total_word_count
        fog_index = 0.4*(avg_sentence_length + perc_complex_words)

        # 3. Average Number of Words Per Sentence
        avg_no_words_per_sentence = total_word_count/sentence_count

        # 5. Personal Pronouns
        count_pronouns = getPersonalPronouns(punt_free_words)
        
        # 6. Syllable Count Per Word
        syllable_counts = countSyllables(punt_free_words)

        # 7. Average Word Length
        avg_word_length = getAvgWordLength(punt_free_words, total_word_count)


        # print(f'positive score {positive_score}')
        # print(f'negative score {negative_score}')
        # print(f'polar score {polar_score}')
        # print(f'subjective score {subjective_score}')
        # print(f'avg sentence len {avg_sentence_length}')
        # print(f'perc complex words {perc_complex_words}')
        # print(f'fog index {fog_index}')
        # print(f'avg no words per sentence {avg_no_words_per_sentence}')
        # print(f'count complex words {count_complex_words}')
        # print(f'word count: {nltk_free_wc}')
        # print(f'syllable counts {syllable_counts}')
        # print(f'count pronouns {count_pronouns}')
        # print(f'avg word length {avg_word_length}')
        # print(f'total word length {total_word_count}')
    

        informations = [positive_score, negative_score, polar_score, subjective_score, avg_sentence_length, perc_complex_words, fog_index, avg_no_words_per_sentence, count_complex_words, nltk_free_wc, syllable_counts, count_pronouns, avg_word_length]
        
        
        for col, info in zip(range(3, 16), informations):
            ws_output[get_column_letter(col) + str(row+2)] = info

        
headings = ['POSITIVE SCORE', 'NEGATIVE SCORE', 'POLARITY SCORE', 'SUBJECTIVITY SCORE', 'AVG SENTENCE LENGTH', 'PERCENTAGE OF COMPLEX WORDS', 'FOG INDEX', 'AVG NUMBER OF WORDS PER SENTENCE', 'COMPLEX WORD COUNT', 'WORD COUNT', 'SYLLABLE PER WORD', 'PERSONAL PRONOUNS', 'AVG WORD LENGTH']

for col, heading in zip(range(3,16), headings):
    ws_output[get_column_letter(col) + '1'] = heading
    
wb_output.save('Output_AnuraagaNath.xlsx')
wb_output.close()
print('Task Completed Sucessfully')

